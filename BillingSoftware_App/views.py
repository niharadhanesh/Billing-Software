from django.shortcuts import render,redirect
from django.contrib.auth import authenticate, login,logout
from BillingSoftware_App.models import User,Modules,Companies,Currency,DateFormat,Tax,CompanyLogo,Template,ExtraCostList,PaymentTerms,Miscellaneous,PredefinedText,InvoiceSettings,OrderSettings,EstimateSettings,PurchaseOrderSettings
from django.contrib import messages
from django.contrib.auth.models import User,auth
from django.core.mail import send_mail
from django.conf import settings
from django.urls import reverse
import json
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.db.models import Max
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from .models import EmailTemplate, Template,Customer,CustomerCategory
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from .models import PaymentSettings

# Create your views here.

def index(request):
    if not PredefinedText.objects.exists():
        create_default_predefined_text()

    if not ExtraCostList.objects.exists():
        create_default_extra_costs()

    if not PaymentTerms.objects.exists():
        create_default_payment_terms()

    if not Miscellaneous.objects.filter(is_default=True).exists():
        create_default_misc_settings()

    try:
        template1 = Template.objects.get(template_name='Professional 1')
    except Template.DoesNotExist:
        template1 = None
        print("Warning: 'Professional 1' template not found.")

    try:
        template2 = Template.objects.get(template_name='Professional 2')
    except Template.DoesNotExist:
        template2 = None
        print("Warning: 'Professional 2' template not found.")

    # Invoice Settings
    if template1 and not InvoiceSettings.objects.filter(template=template1).exists():
        create_default_invoice_settings(template1)
    if template2 and not InvoiceSettings.objects.filter(template=template2).exists():
        create_default_invoice_settings(template2)

    # Order Settings
    if template1 and not OrderSettings.objects.filter(template=template1).exists():
        create_default_order_settings(template1)
    if template2 and not OrderSettings.objects.filter(template=template2).exists():
        create_default_order_settings(template2)

    # Estimate Settings
    if template1 and not EstimateSettings.objects.filter(template_name=template1).exists():
        create_default_estimate_settings(template1)
    if template2 and not EstimateSettings.objects.filter(template_name=template2).exists():
        create_default_estimate_settings(template2)

    # Purchase Order Settings
    if template1 and not PurchaseOrderSettings.objects.filter(template_name=template1).exists():
        create_default_purchase_order_settings(template1)
    if template2 and not PurchaseOrderSettings.objects.filter(template_name=template2).exists():
        create_default_purchase_order_settings(template2)

    return render(request, 'index.html')




def Log_in(request):
    return render(request,'login.html')


def adminDashboard(request):
    user = request.user
    return render(request,'admin/admin_dashboard.html',{'user':user,})


def Fun_login(request):
    if request.method == 'POST':
        username=request.POST['username']
        password=request.POST['pswd']
        user1=auth.authenticate(username=username,password=password)
        if user1 is not None:
            # request.session["uid"] = user1.id
            if user1.is_staff:
                login(request,user1)
                return redirect('adminDashboard')
            else:
                auth.login(request,user1)
                return redirect('userDashboard')
        else:
            messages.info(request,'Invalid Username or Password')
            return redirect('Log_in')
    return render(request,'login.html')



def admin_settings(request):
    users = Modules.objects.exclude(user__is_superuser=True).select_related('user').order_by('-id')
    for u in users:
            print("USER:", u.user.username, "| EMAIL:", u.user.email, "| PERMS:", u.create_invoice, u.delete_invoice)
    return render(request,'admin/admin_settings.html',{'users':users})


def LogOut(request):
    auth.logout(request)
    return redirect('index')



def add_user(request):
    if request.method == 'POST':
        name = request.POST['name']
        email = request.POST['email']
        password = request.POST['pswd']
        confirm_password = request.POST['cpswd']

        if password != confirm_password:
            return redirect('admin_settings')

        user = User(username=name, email=email)
        user.set_password(password)
        user.save()

        Modules.objects.create(user=user)

        subject = "Your Account Details"
        message = f"""Dear {name},

            Your account has been created successfully.

            Username: {name}
            Password: {password}
            You can now log in and start using the system.

            Regards,
            Billing Software"""

        send_mail(subject, message, settings.EMAIL_HOST_USER, [email])

        return redirect(f"{reverse('admin_settings')}#admin")
    return render(request,'admin/admin_settings.html')



from django.core.mail import send_mail
from django.conf import settings

def update_user_profile(request, user_id):
    if request.method == 'POST':
        data = json.loads(request.body)

        try:
            user = User.objects.get(id=user_id)
            old_password = user.password 
            user.username = data.get('username', user.username)
            user.email = data.get('email', user.email)

            password_updated = False
            if data.get('password'):
                user.set_password(data['password'])
                password_updated = True

            user.save()

            permissions = data.get('permissions', {})
            modules = Modules.objects.get(user=user)

            for key, val in permissions.items():
                if hasattr(modules, key):
                    setattr(modules, key, bool(val))

            modules.save()

            if password_updated:
                subject = "Your Account Has Been Updated"
                message = f"""Dear {user.username},

                Your account has been updated successfully.

                Username: {user.username}
                Password: {data['password']}
                You can now log in using your new credentials.

                Regards,
                Billing Software
                """
                send_mail(subject, message, settings.EMAIL_HOST_USER, [user.email])

            return JsonResponse({'success': True})

        except Exception as e:
            print("Error updating user:", e)
            return JsonResponse({'success': False})
        


def delete_user(request, user_id):
    if request.method == 'POST':
        try:
            user = User.objects.get(id=user_id)

            Modules.objects.filter(user=user).delete()

            user.delete()

            return JsonResponse({'success': True})

        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'User not found'})

    return JsonResponse({'success': False, 'error': 'Invalid request'})


def company_settings(request):
    company = Companies.objects.first()
    currency = Currency.objects.filter(company=company).first()
    date_format = DateFormat.objects.filter(company=company).first()
    tax = Tax.objects.filter(company=company).first()
    logo = CompanyLogo.objects.filter(company=company).first()

    context = {
        'company': company,
        'currency': currency,
        'date_format': date_format,
        'tax': tax,
        'logo': logo,
        'currency_list': ['USD', 'US', 'EUR', 'GBP', 'INR', 'CAD', 'AUD', 'JPN'],
        'currency_sign_list': ['$', '€', '£', '₹', 'C$', 'A$', '¥'],
        'placement_list': ['Before amount', 'After amount', 'Before amount with space', 'After amount with space'],
        'separator_list': ['.', ','],
        'date_format_list': [
            'dd/mm/yyyy', 'mm/dd/yyyy', 'yyyy/mm/dd',
            'yyyy-mm-dd', 'mm-dd-yyyy', 'dd-mm-yyyy',
            'yyyy.mm.dd', 'dd.mm.yyyy'
        ],
    }

    return render(request, 'admin/company_settings.html', context)



def add_company(request):
    if request.method == 'POST':
        company_name = request.POST.get('company_name')
        address = request.POST.get('company_address')
        email = request.POST.get('company_email')
        sales_tax = request.POST.get('sales_tax')

        currency_name = request.POST.get('currency_name')
        currency_sign = request.POST.get('currency_sign')
        currency_sign_placement = request.POST.get('currency_sign_placement')
        decimal_separator = request.POST.get('decimal_separator')
        date_format = request.POST.get('date_format')

        tax_type_first = request.POST.get('taxType')  
        tax1_name = request.POST.get('tax1_name')
        tax1_rate = request.POST.get('tax1_rate') or 0
        tax2_name = request.POST.get('tax2_name')
        tax2_rate = request.POST.get('tax2_rate') or 0
        print_level_1 = True if request.POST.get('print_level_1') else False
        print_level_2 = True if request.POST.get('print_level_2') else False
        tax_type_second = request.POST.get('tax2_based_on') 

        logo = request.FILES.get('company_logo')
        print_logo = True if request.POST.get('print_logo') else False

        company = Companies.objects.create(
            company_name=company_name,
            address=address,
            email=email,
            sales_tax_reg_no=sales_tax
        )

        Currency.objects.create(
            company=company,
            currency_name=currency_name,
            currency_sign=currency_sign,
            currency_sign_placement=currency_sign_placement,
            decimal_separator=decimal_separator
        )

        DateFormat.objects.create(
            company=company,
            format=date_format
        )

        Tax.objects.create(
            company=company,
            tax_type_first=tax_type_first,
            tax1_name=tax1_name,
            tax_rate=tax1_rate,
            tax2_name=tax2_name,
            tax2_rate=tax2_rate,
            print_tax1=print_level_1,
            print_tax2=print_level_2,
            tax_type_second=tax_type_second
        )

        CompanyLogo.objects.create(
            company=company,
            image=logo,
            print=print_logo
        )

        return redirect('company_settings')  

    return render(request,'admin/company_settings.html')



def edit_company(request, company_id):
    company = get_object_or_404(Companies, id=company_id)

    if request.method == 'POST':
        company.company_name = request.POST.get('company_name') or company.company_name
        company.address = request.POST.get('company_address') or company.address
        company.email = request.POST.get('company_email') or company.email
        company.sales_tax_reg_no = request.POST.get('sales_tax') or company.sales_tax_reg_no
        company.save()

        currency, _ = Currency.objects.get_or_create(company=company)
        currency.currency_name = request.POST.get('currency_name') or currency.currency_name
        currency.currency_sign = request.POST.get('currency_sign') or currency.currency_sign
        currency.currency_sign_placement = request.POST.get('currency_sign_placement') or currency.currency_sign_placement
        currency.decimal_separator = request.POST.get('decimal_separator') or currency.decimal_separator
        currency.save()

        date_format, _ = DateFormat.objects.get_or_create(company=company)
        date_format.format = request.POST.get('date_format') or date_format.format
        date_format.save()

        tax, _ = Tax.objects.get_or_create(company=company)
        tax.tax_type_first = request.POST.get('taxType') or tax.tax_type_first
        tax.tax1_name = request.POST.get('tax1_name') or tax.tax1_name
        tax.tax_rate = request.POST.get('tax1_rate') or tax.tax_rate
        tax.tax2_name = request.POST.get('tax2_name') or tax.tax2_name
        tax.tax2_rate = request.POST.get('tax2_rate') or tax.tax2_rate
        tax.print_tax1 = bool(request.POST.get('print_level_1'))
        tax.print_tax2 = bool(request.POST.get('print_level_2'))
        tax.tax_type_second = request.POST.get('tax2_based_on') or tax.tax_type_second
        tax.save()

        logo_file = request.FILES.get('company_logo')
        print_logo = bool(request.POST.get('print_logo'))

        logo_obj, _ = CompanyLogo.objects.get_or_create(company=company)
        if logo_file:
            logo_obj.image = logo_file
        logo_obj.print = print_logo
        logo_obj.save()

        next_url = request.POST.get('next')
        if next_url:
            return redirect(next_url)

        return redirect('company_settings')  

    return render(request, 'admin/company_settings.html')


def advanced_settings(request):
    company = Companies.objects.first()
    logo = CompanyLogo.objects.filter(company=company, print=True).first()
    template1_obj = None
    template2_obj = None

    try:
        template1_obj = Template.objects.get(template_name='Professional 1')
    except Template.DoesNotExist:
        print("Warning: 'Professional 1' template not found in database.")

    try:
        template2_obj = Template.objects.get(template_name='Professional 2')
    except Template.DoesNotExist:
        print("Warning: 'Professional 2' template not found in database.")


    return render(request, 'admin/advanced_settings.html', {
        'company': company,
        'logo': logo,
        'template1_obj': template1_obj,
        'template2_obj': template2_obj,
    })




def advanced_template_settings(request):
    if request.method == "POST":
        for i in [1, 2]:
            name = f"Professional {i}"
            page_size = request.POST.get(f"template{i}_page_size")
            right_margin = request.POST.get(f"template{i}_right_margin") or 0
            shift_left = request.POST.get(f"template{i}_shift_left") or 0
            shift_top = request.POST.get(f"template{i}_shift_top") or 0

            template_obj, created = Template.objects.get_or_create(template_name=name)
            template_obj.page_size = page_size
            template_obj.right_margin_mm = right_margin
            template_obj.invoice_block_shift_left_mm = shift_left
            template_obj.invoice_block_shift_top_mm = shift_top
            template_obj.save()

        return redirect('advanced_settings')  

    templates = Template.objects.all()
    template_data = {t.template_name: t for t in templates}

    return render(request, 'admin/advanced_settings.html', {'template_data': template_data})




def restore_template_defaults(request):
    if request.method == "POST":
        template_name = request.POST.get("template_name", "").strip()

        if not template_name or template_name.lower() == "professional 1":
            try:
                template1 = Template.objects.get(template_name__iexact="Professional 1")
                template1.page_size = "A4"
                template1.right_margin_mm = 0
                template1.invoice_block_shift_left_mm = 0
                template1.invoice_block_shift_top_mm = 0
                template1.save()
                template1.refresh_from_db()
            except Template.DoesNotExist:
                pass
            except Exception as e:
                import traceback
                traceback.print_exc()

        if not template_name or template_name.lower() == "professional 2":
            try:
                template2 = Template.objects.get(template_name__iexact="Professional 2")
                template2.page_size = "A4"
                template2.right_margin_mm = 0
                template2.invoice_block_shift_left_mm = 0
                template2.invoice_block_shift_top_mm = 0
                template2.save()
                template2.refresh_from_db()
            except Template.DoesNotExist:
                pass
            except Exception as e:
                import traceback
                traceback.print_exc()

        return redirect('advanced_settings')

    return redirect('advanced_settings')



def userDashboard(request):
    return render(request,'user/user_dashboard.html')


def user_settings(request):
    company = Companies.objects.first()
    currency = Currency.objects.filter(company=company).first()
    date_format = DateFormat.objects.filter(company=company).first()
    tax = Tax.objects.filter(company=company).first()
    logo = CompanyLogo.objects.filter(company=company).first()

    context = {
        'company': company,
        'currency': currency,
        'date_format': date_format,
        'tax': tax,
        'logo': logo,
        'currency_list': ['USD', 'US', 'EUR', 'GBP', 'INR', 'CAD', 'AUD', 'JPN'],
        'currency_sign_list': ['$', '€', '£', '₹', 'C$', 'A$', '¥'],
        'placement_list': ['Before amount', 'After amount', 'Before amount with space', 'After amount with space'],
        'separator_list': ['.', ','],
        'date_format_list': [
            'dd/mm/yyyy', 'mm/dd/yyyy', 'yyyy/mm/dd',
            'yyyy-mm-dd', 'mm-dd-yyyy', 'dd-mm-yyyy',
            'yyyy.mm.dd', 'dd.mm.yyyy'
        ],
    }

    return render(request, 'user/user_settings.html', context)




def user_advanced_settings(request):
    company = Companies.objects.first()
    logo = CompanyLogo.objects.filter(company=company, print=True).first()
    template1_obj = None
    template2_obj = None

    try:
        template1_obj = Template.objects.get(template_name='Professional 1')
    except Template.DoesNotExist:
        print("Warning: 'Professional 1' template not found in database.")

    try:
        template2_obj = Template.objects.get(template_name='Professional 2')
    except Template.DoesNotExist:
        print("Warning: 'Professional 2' template not found in database.")


    return render(request, 'user/advanced_settings.html', {
        'company': company,
        'logo': logo,
        'template1_obj': template1_obj,
        'template2_obj': template2_obj,
    })



def user_add_company(request):
    if request.method == 'POST':
        company_name = request.POST.get('company_name')
        address = request.POST.get('company_address')
        email = request.POST.get('company_email')
        sales_tax = request.POST.get('sales_tax')

        currency_name = request.POST.get('currency_name')
        currency_sign = request.POST.get('currency_sign')
        currency_sign_placement = request.POST.get('currency_sign_placement')
        decimal_separator = request.POST.get('decimal_separator')
        date_format = request.POST.get('date_format')

        tax_type_first = request.POST.get('taxType')  
        tax1_name = request.POST.get('tax1_name')
        tax1_rate = request.POST.get('tax1_rate') or 0
        tax2_name = request.POST.get('tax2_name')
        tax2_rate = request.POST.get('tax2_rate') or 0
        print_level_1 = True if request.POST.get('print_level_1') else False
        print_level_2 = True if request.POST.get('print_level_2') else False
        tax_type_second = request.POST.get('tax2_based_on') 

        logo = request.FILES.get('company_logo')
        print_logo = True if request.POST.get('print_logo') else False

        company = Companies.objects.create(
            company_name=company_name,
            address=address,
            email=email,
            sales_tax_reg_no=sales_tax
        )

        Currency.objects.create(
            company=company,
            currency_name=currency_name,
            currency_sign=currency_sign,
            currency_sign_placement=currency_sign_placement,
            decimal_separator=decimal_separator
        )

        DateFormat.objects.create(
            company=company,
            format=date_format
        )

        Tax.objects.create(
            company=company,
            tax_type_first=tax_type_first,
            tax1_name=tax1_name,
            tax_rate=tax1_rate,
            tax2_name=tax2_name,
            tax2_rate=tax2_rate,
            print_tax1=print_level_1,
            print_tax2=print_level_2,
            tax_type_second=tax_type_second
        )

        CompanyLogo.objects.create(
            company=company,
            image=logo,
            print=print_logo
        )

        return redirect('/user_settings?open=company')

    return render(request,'user/user_settings.html')


def user_edit_company(request, company_id):
    company = get_object_or_404(Companies, id=company_id)

    if request.method == 'POST':
        company.company_name = request.POST.get('company_name') or company.company_name
        company.address = request.POST.get('company_address') or company.address
        company.email = request.POST.get('company_email') or company.email
        company.sales_tax_reg_no = request.POST.get('sales_tax') or company.sales_tax_reg_no
        company.save()

        currency, _ = Currency.objects.get_or_create(company=company)
        currency.currency_name = request.POST.get('currency_name') or currency.currency_name
        currency.currency_sign = request.POST.get('currency_sign') or currency.currency_sign
        currency.currency_sign_placement = request.POST.get('currency_sign_placement') or currency.currency_sign_placement
        currency.decimal_separator = request.POST.get('decimal_separator') or currency.decimal_separator
        currency.save()

        date_format, _ = DateFormat.objects.get_or_create(company=company)
        date_format.format = request.POST.get('date_format') or date_format.format
        date_format.save()

        tax, _ = Tax.objects.get_or_create(company=company)
        tax.tax_type_first = request.POST.get('taxType') or tax.tax_type_first
        tax.tax1_name = request.POST.get('tax1_name') or tax.tax1_name
        tax.tax_rate = request.POST.get('tax1_rate') or tax.tax_rate
        tax.tax2_name = request.POST.get('tax2_name') or tax.tax2_name
        tax.tax2_rate = request.POST.get('tax2_rate') or tax.tax2_rate
        tax.print_tax1 = True if request.POST.get('print_level_1') else False
        tax.print_tax2 = True if request.POST.get('print_level_2') else False
        tax.tax_type_second = request.POST.get('tax2_based_on') or tax.tax_type_second
        tax.save()

        logo_file = request.FILES.get('company_logo')
        print_logo = True if request.POST.get('print_logo') else False

        logo_obj, _ = CompanyLogo.objects.get_or_create(company=company)
        if logo_file:
            logo_obj.image = logo_file
        logo_obj.print = print_logo
        logo_obj.save()

        return redirect('/user_settings?open=company')

    return render(request, 'user/user_settings.html')



def user_advanced_template_settings(request):
    if request.method == "POST":
        for i in [1, 2]:
            name = f"Professional {i}"
            page_size = request.POST.get(f"template{i}_page_size")
            right_margin = request.POST.get(f"template{i}_right_margin") or 0
            shift_left = request.POST.get(f"template{i}_shift_left") or 0
            shift_top = request.POST.get(f"template{i}_shift_top") or 0

            template_obj, created = Template.objects.get_or_create(template_name=name)
            template_obj.page_size = page_size
            template_obj.right_margin_mm = right_margin
            template_obj.invoice_block_shift_left_mm = shift_left
            template_obj.invoice_block_shift_top_mm = shift_top
            template_obj.save()

        return redirect('user_advanced_settings')  

    templates = Template.objects.all()
    template_data = {t.template_name: t for t in templates}

    return render(request, 'user/advanced_settings.html', {'template_data': template_data})




def user_restore_template_defaults(request):
    if request.method == "POST":
        template_name = request.POST.get("template_name", "").strip()

        if not template_name or template_name.lower() == "professional 1":
            try:
                template1 = Template.objects.get(template_name__iexact="Professional 1")
                template1.page_size = "A4"
                template1.right_margin_mm = 0
                template1.invoice_block_shift_left_mm = 0
                template1.invoice_block_shift_top_mm = 0
                template1.save()
                template1.refresh_from_db()
            except Template.DoesNotExist:
                pass
            except Exception as e:
                import traceback
                traceback.print_exc()

        if not template_name or template_name.lower() == "professional 2":
            try:
                template2 = Template.objects.get(template_name__iexact="Professional 2")
                template2.page_size = "A4"
                template2.right_margin_mm = 0
                template2.invoice_block_shift_left_mm = 0
                template2.invoice_block_shift_top_mm = 0
                template2.save()
                template2.refresh_from_db()
            except Template.DoesNotExist:
                pass
            except Exception as e:
                import traceback
                traceback.print_exc()

        return redirect('user_advanced_settings')

    return redirect('user_advanced_settings')



# ------------------------------------------------SECOND PART-------------------------------------------


def miscellaneous_settings(request):
    costs = ExtraCostList.objects.all()
    predefined_texts = PredefinedText.objects.all()
    terms = PaymentTerms.objects.all()

    misc = Miscellaneous.objects.filter(is_default=True).first()
    confirm_logout = misc.confirmation if misc else False
    saved_theme = misc.menu_color if misc else "win7"

    return render(request, 'admin/miscellaneous.html', {
        'costs': costs,
        'texts': predefined_texts,
        'terms': terms,
        'misc': misc,
        'confirm_logout': confirm_logout,
        'saved_theme': saved_theme,
        'update_misc_url': reverse('update_misc_settings'),
    })


def create_default_extra_costs():
    default_costs = ["Shipping and handling", "Postage and handling ", "Delivery cost "]
    for cost in default_costs:
        ExtraCostList.objects.get_or_create(cost_name=cost)


def add_new_cost(request):
    if request.method == 'POST':
        cost = request.POST.get('extraCostName')
        if cost: 
            ExtraCostList.objects.create(cost_name=cost)
        return redirect('miscellaneous_settings')
    return redirect('miscellaneous_settings')


def edit_cost(request, cost_id):
    if request.method == 'POST':
        cost = get_object_or_404(ExtraCostList, id=cost_id)
        costname = request.POST.get('extraCostName')
        if costname:  
            cost.cost_name = costname
            cost.save()
        return redirect('miscellaneous_settings')
    return redirect('miscellaneous_settings')


def delete_cost(request, cost_id):
    if request.method == 'POST':
        cost = get_object_or_404(ExtraCostList, id=cost_id)
        cost.delete()
        return redirect('miscellaneous_settings')



def create_default_predefined_text():
    default_predefined_texts = [
        "Thank you for your purchase!",
        "Thank you for buying!",
        "Thank you for your business!",
        "Thank you for your order!"
    ]
    for text in default_predefined_texts:
        PredefinedText.objects.get_or_create(predefined_text=text)


def add_new_predefined_text(request):
    if request.method == 'POST':
        text = request.POST.get('headerFooterInput')
        if text: 
            PredefinedText.objects.create(predefined_text=text)
        return redirect('miscellaneous_settings')
    return redirect('miscellaneous_settings')


def edit_header_footer(request, text_id):
    text_obj = get_object_or_404(PredefinedText, id=text_id)
    if request.method == 'POST':
        new_text = request.POST.get("headerFootertext")
        if new_text:
            text_obj.predefined_text = new_text
            text_obj.save()
    return redirect('miscellaneous_settings')


def delete_header_footer(request, text_id):
    text_obj = get_object_or_404(PredefinedText, id=text_id)
    text_obj.delete()
    return redirect('miscellaneous_settings')


def validate_old_password(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        password = data.get('password')
        user = request.user
        is_valid = user.check_password(password)
        return JsonResponse({'valid': is_valid})
    


def change_password(request):
    if request.method == 'POST':
        try:
            print("Raw body:", request.body)
            data = json.loads(request.body.decode('utf-8'))
            print("Parsed data:", data)
        except json.JSONDecodeError:
            print("JSON error")
            return JsonResponse({'success': False, 'error': 'Invalid JSON'})

        old_password = data.get('old_password')
        new_password = data.get('new_password')
        confirm_password = data.get('confirm_password')

        print("Old:", old_password)
        print("New:", new_password)
        print("Confirm:", confirm_password)

        if not old_password or not new_password or not confirm_password:
            return JsonResponse({'success': False, 'error': 'All fields are required.'})

        user = request.user
        print("User:", user.username)

        if not user.is_authenticated:
            print("Not authenticated")
            return JsonResponse({'success': False, 'error': 'User not authenticated'})

        if not user.check_password(old_password):
            print("Old password incorrect")
            return JsonResponse({'success': False, 'error': 'Old password is incorrect.'})

        if new_password != confirm_password:
            print("Passwords don't match")
            return JsonResponse({'success': False, 'error': 'Passwords do not match.'})

        print("Changing password for", user.username)
        user.set_password(new_password)
        user.save()

        logout(request)
        print("Logged out")
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})



def create_default_payment_terms():
    default_terms = [
        ("Cash on delivery", 0),
        ("Credit card", 0),
        ("NET 07", 7),
        ("NET 10", 10),
        ("NET 10", 14),
        ("NET 10", 20),
        ("NET 10", 30),
    ]

    for term_name, days in default_terms:
        PaymentTerms.objects.get_or_create(term_name=term_name, days=days)


def add_new_payment_terms(request):
    if request.method == 'POST':
        terms = request.POST.get('termsName')
        days = request.POST.get('shiftDays')

        if terms and days:
            PaymentTerms.objects.create(term_name=terms, days=int(days))

        return redirect('miscellaneous_settings')
    
    return redirect('miscellaneous_settings')



def edit_payment_term(request, term_id):
    if request.method == 'POST':
        term = get_object_or_404(PaymentTerms, id=term_id)
        name = request.POST.get('termsEditName')
        days = request.POST.get('shitfEditDays')
        if name and days is not None:
            term.term_name = name
            term.days = int(days)
            term.save()
    return redirect('miscellaneous_settings')


def delete_payment_term(request, term_id):
    PaymentTerms.objects.filter(id=term_id).delete()
    return redirect('miscellaneous_settings')



def create_default_misc_settings():
    if not Miscellaneous.objects.filter(is_default=True).exists():
        Miscellaneous.objects.create(
            menu_color='win7',
            attachment_type='pdf',
            invoice_numbering=True,
            order_numbering=True,
            estimate_numbering=True,
            porder_numbering=True,
            confirmation=True,
            is_default=True  
        )


def update_misc_settings(request):
    if request.method == "POST":
        misc = Miscellaneous.objects.filter(is_default=True).first()

        if misc:
            misc.menu_color = request.POST.get('menu_color', misc.menu_color)
            misc.attachment_type = request.POST.get('attachment_type', misc.attachment_type)
            misc.invoice_numbering = 'invoice_numbering' in request.POST
            misc.order_numbering = 'order_numbering' in request.POST
            misc.estimate_numbering = 'estimate_numbering' in request.POST
            misc.porder_numbering = 'porder_numbering' in request.POST
            misc.confirmation = 'confirmation' in request.POST
            misc.save()
            return JsonResponse({'success': True, 'message': 'Settings updated successfully.'})
        else:
            return JsonResponse({'success': False, 'error': 'Default miscellaneous settings not found.'})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})



def invoice_settings(request):
    company = Companies.objects.first()
    templates = Template.objects.all()
    logo = CompanyLogo.objects.filter(company=company, print=True).first()

    template1_obj = Template.objects.filter(template_name='Professional 1').first()
    template2_obj = Template.objects.filter(template_name='Professional 2').first()

    template1_settings = InvoiceSettings.objects.filter(template=template1_obj).first()
    template2_settings = InvoiceSettings.objects.filter(template=template2_obj).first()

    template_id = request.GET.get('template')
    template = Template.objects.filter(id=template_id).first() if template_id else None
    active_settings = InvoiceSettings.objects.filter(template=template).first() if template else None

    return render(request, 'admin/invoice_settings.html', {
        'company': company,
        'logo': logo,
        'template1_obj': template1_obj,
        'template2_obj': template2_obj,
        'templates': templates,
        'template1_settings': template1_settings,
        'template2_settings': template2_settings,
        'active_settings': active_settings,
        'selected_template': template,
    })



def create_default_invoice_settings(template):
    InvoiceSettings.objects.get_or_create(
        template=template,
        defaults={
            'invoice_number_prefix': 'INV',
            'invoice_starting_number': 1,
            'header_background_color': '#ffa500',
            'invoice_label': 'Invoice To',
            'invoice_number_label': 'Invoice #',
            'invoice_date_label': 'Invoice Date',
            'due_date_label': 'Due Date',
            'order_ref_label': 'Order #',
            'terms_label': 'Terms',
            'invoice_to_label': 'Invoice To',
            'ship_to_label': 'Ship To',
            'id_sku_label': 'ID/SKU',
            'product_service_label': 'Product/Service',
            'quantity_label': 'Quantity',
            'description_label': 'Description',
            'unit_price_label': 'Unit Price',
            'price_label': 'Price',
            'subtotal_label': 'Subtotal',
            'discount_label': 'Discount',
            'discount_rate': 'Discount Rate',
            'tax1': 'Tax 1',
            'tax2': 'Tax 2',
            'invoice_total': 'Invoice Total',
            'total_paid': 'Total Paid',
            'balance': 'Balance',
            'terms_and_conditions': 'Terms and Conditions',
            'tax_exempted': 'Tax Exempted',
            'page': 'Page',
            'of': 'Of',
            'invoice_terms_footer': 'Invoice Terms Footer'
        }
    )



def update_invoice_settings(request):
    if request.method != 'POST':
        return redirect('invoice_settings')

    section = request.POST.get('section', 'invoice')
    template_id = request.POST.get('template')
    template = get_object_or_404(Template, id=template_id)

    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    settings1, _ = InvoiceSettings.objects.get_or_create(template=template1)
    settings2, _ = InvoiceSettings.objects.get_or_create(template=template2)

    def update_settings(settings):
        if 'invoice_number_prefix1' in request.POST:
            settings.invoice_number_prefix = request.POST['invoice_number_prefix1'].strip()
        if 'invoice_starting_number1' in request.POST:
            settings.invoice_starting_number = request.POST['invoice_starting_number1'] or settings.invoice_starting_number
        if 'header_background_color_final' in request.POST:
            settings.header_background_color = request.POST['header_background_color_final']

        fields = [
            'invoice_label', 'invoice_number_label', 'invoice_date_label',
            'due_date_label', 'order_ref_label', 'terms_label',  'invoice_to_label', 'ship_to_label',
            'id_sku_label', 'product_service_label', 'quantity_label', 'description_label',
            'unit_price_label', 'price_label', 'subtotal_label', 'discount_label',
            'terms_and_conditions', 'tax_exempted', 'page', 'of', 'invoice_terms_footer'
        ]

        for field in fields:
            if field in request.POST:
                setattr(settings, field, request.POST.get(field))

        numeric_fields = ['discount_rate', 'tax1', 'tax2', 'invoice_total', 'total_paid', 'balance']
        for field in numeric_fields:
            if field in request.POST and request.POST.get(field) != "":
                setattr(settings, field, request.POST.get(field))

        settings.save()

    update_settings(settings1)
    update_settings(settings2)

    return redirect(f'/invoice_settings/?section={section}')




def restore_invoice_defaults(request):
    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    default_labels = {
        'invoice_label': "Invoice",
        'invoice_number_label': "Invoice #",
        'invoice_date_label': "Invoice Date",
        'due_date_label': "Due Date",
        'order_ref_label': "Order #",
        'terms_label': "Terms",
        'invoice_to_label': "Invoice To",
        'ship_to_label': "Ship To",
        'id_sku_label': "ID/SKU",
        'product_service_label': "Product/Service",
        'quantity_label': "Quantity",
        'description_label': "Description",
        'unit_price_label': "Unit Price",
        'price_label': "Price",
        'subtotal_label': "Subtotal",
        'discount_label': "Discount",
        'discount_rate': "Discount Rate",   
        'tax1': "Tax 1",                    
        'tax2': "Tax 2",                    
        'invoice_total': "Invoice Total",   
        'total_paid': "Total Paid",         
        'balance': "Balance",               
        'terms_and_conditions': "Terms and Conditions",
        'tax_exempted': "Tax Exempted",
        'page': "Page",
        'of': "of",
        'invoice_terms_footer': ""
    }

    def apply_defaults(template):
        if template:
            settings, _ = InvoiceSettings.objects.get_or_create(template=template)
            settings.invoice_number_prefix = "INV"
            settings.invoice_starting_number = 1
            settings.header_background_color = "#ffa500"
            for key, value in default_labels.items():
                setattr(settings, key, value)
            settings.save()

    apply_defaults(template1)
    apply_defaults(template2)

    section = request.GET.get('section', 'invoice')
    return redirect(f'/invoice_settings/?section={section}')



def order_settings(request):
    company = Companies.objects.first()
    templates = Template.objects.all()
    logo = CompanyLogo.objects.filter(company=company, print=True).first()
    template1_obj = None
    template2_obj = None

    try:
        template1_obj = Template.objects.get(template_name='Professional 1')
    except Template.DoesNotExist:
        print("Warning: 'Professional 1' template not found in database.")

    try:
        template2_obj = Template.objects.get(template_name='Professional 2')
    except Template.DoesNotExist:
        print("Warning: 'Professional 2' template not found in database.")

    if template1_obj:
        create_default_order_settings(template1_obj)
    if template2_obj:
        create_default_order_settings(template2_obj)

    template1_settings = OrderSettings.objects.filter(template=template1_obj).first()
    template2_settings = OrderSettings.objects.filter(template=template2_obj).first()

    template_id = request.GET.get('template')
    template = None
    if template_id:
        try:
            template = Template.objects.get(id=template_id)
        except Template.DoesNotExist:
            print(f"Warning: Template with ID {template_id} not found.")

    active_settings = OrderSettings.objects.filter(template=template).first() if template else None

    return render(request, 'admin/order_settings.html', {
        'company': company,
        'logo': logo,
        'template1_obj': template1_obj,
        'template2_obj': template2_obj,
        'templates': templates,
        'template1_settings': template1_settings,
        'template2_settings': template2_settings,
        'active_settings': active_settings,
        'selected_template': template,
    })



def create_default_order_settings(template):
    OrderSettings.objects.get_or_create(
        template=template,
        defaults={
            'order_number_prefix': 'ORD',
            'order_starting_number': 1,
            'header_background_color': '#ffa500',

            'order_label': 'Order',
            'order_number_label': 'Order #',
            'order_date_label': 'Order Date',
            'due_date_label': 'Due Date',
            'order_to': 'Order To',
            'order_total': 'Order Total',
            'order_terms_footer': ''
        }
    )


def update_order_settings(request):
    if request.method != 'POST':
        return redirect('order_settings')

    section = request.POST.get('section', 'order')
    template_id = request.POST.get('template')
    template = get_object_or_404(Template, id=template_id)

    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    settings1, _ = OrderSettings.objects.get_or_create(template=template1)
    settings2, _ = OrderSettings.objects.get_or_create(template=template2)

    def update_settings(settings):
        if 'order_number_prefix1' in request.POST:
            settings.order_number_prefix = request.POST['order_number_prefix1'].strip()

        if 'order_starting_number1' in request.POST:
            settings.order_starting_number = request.POST['order_starting_number1'] or settings.order_starting_number

        if 'header_background_color_final' in request.POST:
            settings.header_background_color = request.POST['header_background_color_final']

        fields = [
            'order_label', 'order_number_label', 'order_date_label',
            'due_date_label', 'order_to', 'order_total', 'order_terms_footer' 
        ]

        for field in fields:
            if field in request.POST:
                setattr(settings, field, request.POST.get(field))

        settings.save()

    update_settings(settings1)
    update_settings(settings2)

    return redirect(f'/order_settings/?section={section}')




def restore_order_defaults(request):
    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    default_labels = {
        'order_label': "Order",
        'order_number_label': "Order #",
        'order_date_label': "Order Date",
        'due_date_label': "Due Date",
        'order_to': "Order To",
        'order_total': "Order Total",
        'order_terms_footer': ""
    }

    def apply_defaults(template):
        if template:
            settings, _ = OrderSettings.objects.get_or_create(template=template)
            settings.order_number_prefix = "ORD"
            settings.order_starting_number = 1
            settings.header_background_color = "#ffa500"
            for key, value in default_labels.items():
                setattr(settings, key, value)
            settings.save()

    apply_defaults(template1)
    apply_defaults(template2)

    section = request.GET.get('section', 'order')
    return redirect(f'/order_settings/?section={section}')



def estimate_settings(request):
    company = Companies.objects.first()
    templates = Template.objects.all()
    logo = CompanyLogo.objects.filter(company=company, print=True).first()

    template1_obj = Template.objects.filter(template_name='Professional 1').first()
    template2_obj = Template.objects.filter(template_name='Professional 2').first()

    template1_settings = EstimateSettings.objects.filter(template_name=template1_obj).first()
    template2_settings = EstimateSettings.objects.filter(template_name=template2_obj).first()

    template_id = request.GET.get('template')
    template = Template.objects.filter(id=template_id).first() if template_id else None
    active_settings = EstimateSettings.objects.filter(template=template).first() if template else None

    return render(request, 'admin/estimate_settings.html', {
        'company': company,
        'logo': logo,
        'template1_obj': template1_obj,
        'template2_obj': template2_obj,
        'templates': templates,
        'template1_settings': template1_settings,
        'template2_settings': template2_settings,
        'active_settings': active_settings,
        'selected_template': template,
    })



def create_default_estimate_settings(template):
    EstimateSettings.objects.get_or_create(
        template_name=template,
        defaults={
            'estimate_number_prefix': 'EST',
            'estimate_starting_number': 1,
            'header_background_color': '#ffa500',

            'estimate_label': 'Estimate',
            'estimate_number_label': 'Estimate #',
            'estimate_date_label': 'Estimate Date',
            'due_date_label': 'Due Date',
            'estimate_to': 'Estimate To',
            'estimate_total': 'Estimate Total',
            'estimate_terms_footer': 'Estimate Terms Footer'
        }
    )



def update_estimate_settings(request):
    if request.method != 'POST':
        return redirect('estimate_settings')

    section = request.POST.get('section', 'estimate')

    template_id = request.POST.get('template')
    template = get_object_or_404(Template, id=template_id)

    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    settings1, _ = EstimateSettings.objects.get_or_create(template_name=template1)
    settings2, _ = EstimateSettings.objects.get_or_create(template_name=template2)

    def update_settings(settings):
        if 'estimate_number_prefix1' in request.POST:
            settings.estimate_number_prefix = request.POST['estimate_number_prefix1'].strip()
        if 'estimate_starting_number1' in request.POST:
            settings.estimate_starting_number = request.POST['estimate_starting_number1'] or settings.estimate_starting_number
        if 'header_background_color_final' in request.POST:
            settings.header_background_color = request.POST['header_background_color_final']

        fields = [
            'estimate_label', 'estimate_number_label', 'estimate_date_label',
            'due_date_label', 'estimate_to', 'estimate_total', 'estimate_terms_footer'
        ]

        for field in fields:
            if field in request.POST:
                setattr(settings, field, request.POST.get(field))

        settings.save()

    update_settings(settings1)
    update_settings(settings2)

    return redirect(f'/estimate_settings/?section={section}')



def restore_estimate_defaults(request):
    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    default_labels = {
        'estimate_label': "Estimate",
        'estimate_number_label': "Estimate #",
        'estimate_date_label': "Estimate Date",
        'due_date_label': "Due Date",
        'estimate_to': "Estimate To",
        'estimate_total': "Estimate Total",
        'estimate_terms_footer': ""
    }

    def apply_defaults(template):
        if template:
            settings, _ = EstimateSettings.objects.get_or_create(template_name=template)
            settings.estimate_number_prefix = "EST"
            settings.estimate_starting_number = 1
            settings.header_background_color = "#ffa500"
            for key, value in default_labels.items():
                setattr(settings, key, value)
            settings.save()

    apply_defaults(template1)
    apply_defaults(template2)

    section = request.GET.get('section', 'estimate')
    return redirect(f'/estimate_settings/?section={section}')



def purchase_order(request):
    company = Companies.objects.first()
    templates = Template.objects.all()
    logo = CompanyLogo.objects.filter(company=company, print=True).first()

    template1_obj = Template.objects.filter(template_name='Professional 1').first()
    template2_obj = Template.objects.filter(template_name='Professional 2').first()

    template1_settings = PurchaseOrderSettings.objects.filter(template_name=template1_obj).first()
    template2_settings = PurchaseOrderSettings.objects.filter(template_name=template2_obj).first()

    template_id = request.GET.get('template')
    template = Template.objects.filter(id=template_id).first() if template_id else None
    active_settings = PurchaseOrderSettings.objects.filter(template=template).first() if template else None

    return render(request, 'admin/purchase_order.html', {
        'company': company,
        'logo': logo,
        'template1_obj': template1_obj,
        'template2_obj': template2_obj,
        'templates': templates,
        'template1_settings': template1_settings,
        'template2_settings': template2_settings,
        'active_settings': active_settings,
        'selected_template': template,
    })



def create_default_purchase_order_settings(template):
    PurchaseOrderSettings.objects.get_or_create(
        template_name=template,
        defaults={
            'porder_number_prefix': 'P.ORD',
            'porder_starting_number': 1,
            'header_background_color': '#ffa500',

            'porder_label': 'Purchase Order',
            'porder_number_label': 'P. Order #',
            'porder_date_label': 'P. Order Date',
            'due_date_label': 'Due Date',
            'porder_to': 'Vendor',
            'delivery_to': 'Delivery To',
            'porder_total': 'Purchase Order Total',
            'porder_terms_footer': 'P. Order Terms Footer'
        }
    )



def update_purchase_order_settings(request):
    if request.method != 'POST':
        return redirect('purchase_order')

    section = request.POST.get('section', 'purchase')

    template_id = request.POST.get('template')
    template = get_object_or_404(Template, id=template_id)

    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    settings1, _ = PurchaseOrderSettings.objects.get_or_create(template_name=template1)
    settings2, _ = PurchaseOrderSettings.objects.get_or_create(template_name=template2)

    def update_settings(settings):
        if 'porder_number_prefix1' in request.POST:
            settings.porder_number_prefix = request.POST['porder_number_prefix1'].strip()
        if 'porder_starting_number1' in request.POST:
            settings.porder_starting_number = request.POST['porder_starting_number1'] or settings.porder_starting_number
        if 'header_background_color_final' in request.POST:
            settings.header_background_color = request.POST['header_background_color_final']

        fields = [
            'porder_label', 'porder_number_label', 'porder_date_label',
            'due_date_label', 'porder_to', 'delivery_to', 'porder_total', 'porder_terms_footer'
        ]

        for field in fields:
            if field in request.POST:
                setattr(settings, field, request.POST.get(field))

        settings.save()

    update_settings(settings1)
    update_settings(settings2)

    return redirect(f'/purchase_order/?section={section}')



def restore_purchase_order_defaults(request):
    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    default_labels = {
        'porder_label': "Purchase Order",
        'porder_number_label': "Purchase Order #",
        'porder_date_label': "Purchase Order Date",
        'due_date_label': "Due Date",
        'porder_to': "Vendor",
        'delivery_to': "Delivery To",
        'porder_total': "Purchase Order Total",
        'porder_terms_footer': ""
    }

    def apply_defaults(template):
        if template:
            settings, _ = PurchaseOrderSettings.objects.get_or_create(template_name=template)
            settings.porder_number_prefix = "P.ORD"
            settings.porder_starting_number = 1
            settings.header_background_color = "#ffa500"
            for key, value in default_labels.items():
                setattr(settings, key, value)
            settings.save()

    apply_defaults(template1)
    apply_defaults(template2)

    section = request.GET.get('section', 'purchase')
    return redirect(f'/purchase_order/?section={section}')




def user_miscellaneous_settings(request):
    costs = ExtraCostList.objects.all()
    predefined_texts = PredefinedText.objects.all()
    terms = PaymentTerms.objects.all()

    misc = Miscellaneous.objects.filter(is_default=True).first()
    confirm_logout = misc.confirmation if misc else False
    saved_theme = misc.menu_color if misc else "win7"


    return render(request, 'user/miscellaneous.html', {
        'costs': costs,
        'texts': predefined_texts,
        'terms': terms,
        'misc': misc,
        'confirm_logout': confirm_logout,
        'saved_theme': saved_theme,
        'user_update_misc_url': reverse('user_update_misc_settings'),
    })


def user_add_new_cost(request):
    if request.method == 'POST':
        cost = request.POST.get('extraCostName')
        if cost: 
            ExtraCostList.objects.create(cost_name=cost)
        return redirect('user_miscellaneous_settings')
    return redirect('user_miscellaneous_settings')


def user_edit_cost(request, cost_id):
    if request.method == 'POST':
        cost = get_object_or_404(ExtraCostList, id=cost_id)
        costname = request.POST.get('extraCostName')
        if costname:  
            cost.cost_name = costname
            cost.save()
        return redirect('user_miscellaneous_settings')
    return redirect('user_miscellaneous_settings')


def user_delete_cost(request, cost_id):
    if request.method == 'POST':
        cost = get_object_or_404(ExtraCostList, id=cost_id)
        cost.delete()
        return redirect('user_miscellaneous_settings')



def user_add_new_predefined_text(request):
    if request.method == 'POST':
        text = request.POST.get('headerFooterInput')
        if text: 
            PredefinedText.objects.create(predefined_text=text)
        return redirect('user_miscellaneous_settings')
    return redirect('user_miscellaneous_settings')


def user_edit_header_footer(request, text_id):
    text_obj = get_object_or_404(PredefinedText, id=text_id)
    if request.method == 'POST':
        new_text = request.POST.get("headerFootertext")
        if new_text:
            text_obj.predefined_text = new_text
            text_obj.save()
    return redirect('user_miscellaneous_settings')


def user_delete_header_footer(request, text_id):
    text_obj = get_object_or_404(PredefinedText, id=text_id)
    text_obj.delete()
    return redirect('user_miscellaneous_settings')



def user_add_new_payment_terms(request):
    if request.method == 'POST':
        terms = request.POST.get('termsName')
        days = request.POST.get('shiftDays')

        if terms and days:
            PaymentTerms.objects.create(term_name=terms, days=int(days))

        return redirect('user_miscellaneous_settings')
    
    return redirect('user_miscellaneous_settings')



def user_edit_payment_term(request, term_id):
    if request.method == 'POST':
        term = get_object_or_404(PaymentTerms, id=term_id)
        name = request.POST.get('termsEditName')
        days = request.POST.get('shitfEditDays')
        if name and days is not None:
            term.term_name = name
            term.days = int(days)
            term.save()
    return redirect('user_miscellaneous_settings')


def user_delete_payment_term(request, term_id):
    PaymentTerms.objects.filter(id=term_id).delete()
    return redirect('user_miscellaneous_settings')



def user_update_misc_settings(request):
    if request.method == "POST":
        misc = Miscellaneous.objects.filter(is_default=True).first()

        if misc:
            misc.menu_color = request.POST.get('menu_color', misc.menu_color)
            misc.attachment_type = request.POST.get('attachment_type', misc.attachment_type)
            misc.invoice_numbering = 'invoice_numbering' in request.POST
            misc.order_numbering = 'order_numbering' in request.POST
            misc.estimate_numbering = 'estimate_numbering' in request.POST
            misc.porder_numbering = 'porder_numbering' in request.POST
            misc.confirmation = 'confirmation' in request.POST
            misc.save()
            return JsonResponse({'success': True, 'message': 'Settings updated successfully.'})
        else:
            return JsonResponse({'success': False, 'error': 'Default miscellaneous settings not found.'})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})




def user_invoice_settings(request):
    company = Companies.objects.first()
    templates = Template.objects.all()
    logo = CompanyLogo.objects.filter(company=company, print=True).first()

    template1_obj = Template.objects.filter(template_name='Professional 1').first()
    template2_obj = Template.objects.filter(template_name='Professional 2').first()

    template1_settings = InvoiceSettings.objects.filter(template=template1_obj).first()
    template2_settings = InvoiceSettings.objects.filter(template=template2_obj).first()

    template_id = request.GET.get('template')
    template = Template.objects.filter(id=template_id).first() if template_id else None
    active_settings = InvoiceSettings.objects.filter(template=template).first() if template else None

    return render(request, 'user/invoice_settings.html', {
        'company': company,
        'logo': logo,
        'template1_obj': template1_obj,
        'template2_obj': template2_obj,
        'templates': templates,
        'template1_settings': template1_settings,
        'template2_settings': template2_settings,
        'active_settings': active_settings,
        'selected_template': template,
    })



def user_update_invoice_settings(request):
    if request.method != 'POST':
        return redirect('user_invoice_settings')

    section = request.POST.get('section', 'invoice')
    template_id = request.POST.get('template')
    template = get_object_or_404(Template, id=template_id)

    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    settings1, _ = InvoiceSettings.objects.get_or_create(template=template1)
    settings2, _ = InvoiceSettings.objects.get_or_create(template=template2)

    def update_settings(settings):
        if 'invoice_number_prefix1' in request.POST:
            settings.invoice_number_prefix = request.POST['invoice_number_prefix1'].strip()
        if 'invoice_starting_number1' in request.POST:
            settings.invoice_starting_number = request.POST['invoice_starting_number1'] or settings.invoice_starting_number
        if 'header_background_color_final' in request.POST:
            settings.header_background_color = request.POST['header_background_color_final']

        fields = [
            'invoice_label', 'invoice_number_label', 'invoice_date_label',
            'due_date_label', 'order_ref_label', 'terms_label',  'invoice_to_label', 'ship_to_label',
            'id_sku_label', 'product_service_label', 'quantity_label', 'description_label',
            'unit_price_label', 'price_label', 'subtotal_label', 'discount_label',
            'terms_and_conditions', 'tax_exempted', 'page', 'of', 'invoice_terms_footer' 
        ]

        for field in fields:
            if field in request.POST:
                setattr(settings, field, request.POST.get(field))

        numeric_fields = ['discount_rate', 'tax1', 'tax2', 'invoice_total', 'total_paid', 'balance']
        for field in numeric_fields:
            if field in request.POST and request.POST.get(field) != "":
                setattr(settings, field, request.POST.get(field))

        settings.save()

    update_settings(settings1)
    update_settings(settings2)

    return redirect(f'/user_invoice_settings/?section={section}')




def user_restore_invoice_defaults(request):
    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    default_labels = {
        'invoice_label': "Invoice",
        'invoice_number_label': "Invoice #",
        'invoice_date_label': "Invoice Date",
        'due_date_label': "Due Date",
        'order_ref_label': "Order #",
        'terms_label': "Terms",
        'invoice_to_label': "Invoice To",
        'ship_to_label': "Ship To",
        'id_sku_label': "ID/SKU",
        'product_service_label': "Product/Service",
        'quantity_label': "Quantity",
        'description_label': "Description",
        'unit_price_label': "Unit Price",
        'price_label': "Price",
        'subtotal_label': "Subtotal",
        'discount_label': "Discount",
        'discount_rate': "Discount Rate",   
        'tax1': "Tax 1",                    
        'tax2': "Tax 2",                    
        'invoice_total': "Invoice Total",   
        'total_paid': "Total Paid",         
        'balance': "Balance",               
        'terms_and_conditions': "Terms and Conditions",
        'tax_exempted': "Tax Exempted",
        'page': "Page",
        'of': "of",
        'invoice_terms_footer': ""
    }

    def apply_defaults(template):
        if template:
            settings, _ = InvoiceSettings.objects.get_or_create(template=template)
            settings.invoice_number_prefix = "INV"
            settings.invoice_starting_number = 1
            settings.header_background_color = "#ffa500"
            for key, value in default_labels.items():
                setattr(settings, key, value)
            settings.save()

    apply_defaults(template1)
    apply_defaults(template2)

    section = request.GET.get('section', 'invoice')
    return redirect(f'/user_invoice_settings/?section={section}')



def user_order_settings(request):
    company = Companies.objects.first()
    templates = Template.objects.all()
    logo = CompanyLogo.objects.filter(company=company, print=True).first()

    template1_obj = Template.objects.filter(template_name='Professional 1').first()
    template2_obj = Template.objects.filter(template_name='Professional 2').first()

    template1_settings = OrderSettings.objects.filter(template=template1_obj).first()
    template2_settings = OrderSettings.objects.filter(template=template2_obj).first()

    template_id = request.GET.get('template')
    template = Template.objects.filter(id=template_id).first() if template_id else None
    active_settings = OrderSettings.objects.filter(template=template).first() if template else None

    return render(request, 'user/order_settings.html', {
        'company': company,
        'logo': logo,
        'template1_obj': template1_obj,
        'template2_obj': template2_obj,
        'templates': templates,
        'template1_settings': template1_settings,
        'template2_settings': template2_settings,
        'active_settings': active_settings,
        'selected_template': template,
    })



def user_update_order_settings(request):
    if request.method != 'POST':
        return redirect('user_order_settings')

    section = request.POST.get('section', 'order')
    template_id = request.POST.get('template')
    template = get_object_or_404(Template, id=template_id)

    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    settings1, _ = OrderSettings.objects.get_or_create(template=template1)
    settings2, _ = OrderSettings.objects.get_or_create(template=template2)

    def update_settings(settings):
        if 'order_number_prefix1' in request.POST:
            settings.order_number_prefix = request.POST['order_number_prefix1'].strip()

        if 'order_starting_number1' in request.POST:
            settings.order_starting_number = request.POST['order_starting_number1'] or settings.order_starting_number

        if 'header_background_color_final' in request.POST:
            settings.header_background_color = request.POST['header_background_color_final']

        fields = [
            'order_label', 'order_number_label', 'order_date_label',
            'due_date_label', 'order_to', 'order_total', 'order_terms_footer'  # include only this
        ]

        for field in fields:
            if field in request.POST:
                setattr(settings, field, request.POST.get(field))

        settings.save()

    update_settings(settings1)
    update_settings(settings2)

    return redirect(f'/user_order_settings/?section={section}')



def user_restore_order_defaults(request):
    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    default_labels = {
        'order_label': "Order",
        'order_number_label': "Order #",
        'order_date_label': "Order Date",
        'due_date_label': "Due Date",
        'order_to': "Order To",
        'order_total': "Order Total",
        'order_terms_footer': ""
    }

    def apply_defaults(template):
        if template:
            settings, _ = OrderSettings.objects.get_or_create(template=template)
            settings.order_number_prefix = "ORD"
            settings.order_starting_number = 1
            settings.header_background_color = "#ffa500"
            for key, value in default_labels.items():
                setattr(settings, key, value)
            settings.save()

    apply_defaults(template1)
    apply_defaults(template2)

    section = request.GET.get('section', 'order')
    return redirect(f'/user_order_settings/?section={section}')



def user_estimate_settings(request):
    company = Companies.objects.first()
    templates = Template.objects.all()
    logo = CompanyLogo.objects.filter(company=company, print=True).first()

    template1_obj = Template.objects.filter(template_name='Professional 1').first()
    template2_obj = Template.objects.filter(template_name='Professional 2').first()

    template1_settings = EstimateSettings.objects.filter(template_name=template1_obj).first()
    template2_settings = EstimateSettings.objects.filter(template_name=template2_obj).first()

    template_id = request.GET.get('template')
    template = Template.objects.filter(id=template_id).first() if template_id else None
    active_settings = EstimateSettings.objects.filter(template=template).first() if template else None

    return render(request, 'user/estimate_settings.html', {
        'company': company,
        'logo': logo,
        'template1_obj': template1_obj,
        'template2_obj': template2_obj,
        'templates': templates,
        'template1_settings': template1_settings,
        'template2_settings': template2_settings,
        'active_settings': active_settings,
        'selected_template': template,
    })



def user_update_estimate_settings(request):
    if request.method != 'POST':
        return redirect('user_estimate_settings')

    section = request.POST.get('section', 'estimate')

    template_id = request.POST.get('template')
    template = get_object_or_404(Template, id=template_id)

    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    settings1, _ = EstimateSettings.objects.get_or_create(template_name=template1)
    settings2, _ = EstimateSettings.objects.get_or_create(template_name=template2)

    def update_settings(settings):
        if 'estimate_number_prefix1' in request.POST:
            settings.estimate_number_prefix = request.POST['estimate_number_prefix1'].strip()
        if 'estimate_starting_number1' in request.POST:
            settings.estimate_starting_number = request.POST['estimate_starting_number1'] or settings.estimate_starting_number
        if 'header_background_color_final' in request.POST:
            settings.header_background_color = request.POST['header_background_color_final']

        fields = [
            'estimate_label', 'estimate_number_label', 'estimate_date_label',
            'due_date_label', 'estimate_to', 'estimate_total', 'estimate_terms_footer'
        ]

        for field in fields:
            if field in request.POST:
                setattr(settings, field, request.POST.get(field))

        settings.save()

    update_settings(settings1)
    update_settings(settings2)

    return redirect(f'/user_estimate_settings/?section={section}')



def user_restore_estimate_defaults(request):
    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    default_labels = {
        'estimate_label': "Estimate",
        'estimate_number_label': "Estimate #",
        'estimate_date_label': "Estimate Date",
        'due_date_label': "Due Date",
        'estimate_to': "Estimate To",
        'estimate_total': "Estimate Total",
        'estimate_terms_footer': ""
    }

    def apply_defaults(template):
        if template:
            settings, _ = EstimateSettings.objects.get_or_create(template_name=template)
            settings.estimate_number_prefix = "EST"
            settings.estimate_starting_number = 1
            settings.header_background_color = "#ffa500"
            for key, value in default_labels.items():
                setattr(settings, key, value)
            settings.save()

    apply_defaults(template1)
    apply_defaults(template2)

    section = request.GET.get('section', 'estimate')
    return redirect(f'/user_estimate_settings/?section={section}')



def user_purchase_order(request):
    company = Companies.objects.first()
    templates = Template.objects.all()
    logo = CompanyLogo.objects.filter(company=company, print=True).first()

    template1_obj = Template.objects.filter(template_name='Professional 1').first()
    template2_obj = Template.objects.filter(template_name='Professional 2').first()

    template1_settings = PurchaseOrderSettings.objects.filter(template_name=template1_obj).first()
    template2_settings = PurchaseOrderSettings.objects.filter(template_name=template2_obj).first()

    template_id = request.GET.get('template')
    template = Template.objects.filter(id=template_id).first() if template_id else None
    active_settings = PurchaseOrderSettings.objects.filter(template=template).first() if template else None

    return render(request, 'user/purchase_order.html', {
        'company': company,
        'logo': logo,
        'template1_obj': template1_obj,
        'template2_obj': template2_obj,
        'templates': templates,
        'template1_settings': template1_settings,
        'template2_settings': template2_settings,
        'active_settings': active_settings,
        'selected_template': template,
    })



def user_update_purchase_order_settings(request):
    if request.method != 'POST':
        return redirect('user_purchase_order')

    section = request.POST.get('section', 'purchase')

    template_id = request.POST.get('template')
    template = get_object_or_404(Template, id=template_id)

    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    settings1, _ = PurchaseOrderSettings.objects.get_or_create(template_name=template1)
    settings2, _ = PurchaseOrderSettings.objects.get_or_create(template_name=template2)

    def update_settings(settings):
        if 'porder_number_prefix1' in request.POST:
            settings.porder_number_prefix = request.POST['porder_number_prefix1'].strip()
        if 'porder_starting_number1' in request.POST:
            settings.porder_starting_number = request.POST['porder_starting_number1'] or settings.porder_starting_number
        if 'header_background_color_final' in request.POST:
            settings.header_background_color = request.POST['header_background_color_final']

        fields = [
            'porder_label', 'porder_number_label', 'porder_date_label',
            'due_date_label', 'porder_to', 'delivery_to', 'porder_total', 'porder_terms_footer'
        ]

        for field in fields:
            if field in request.POST:
                setattr(settings, field, request.POST.get(field))

        settings.save()

    update_settings(settings1)
    update_settings(settings2)

    return redirect(f'/user_purchase_order/?section={section}')



def user_restore_purchase_order_defaults(request):
    template1 = Template.objects.filter(template_name='Professional 1').first()
    template2 = Template.objects.filter(template_name='Professional 2').first()

    default_labels = {
        'porder_label': "Purchase Order",
        'porder_number_label': "Purchase Order #",
        'porder_date_label': "Purchase Order Date",
        'due_date_label': "Due Date",
        'porder_to': "Vendor",
        'delivery_to': "Delivery To",
        'porder_total': "Purchase Order Total",
        'porder_terms_footer': ""
    }

    def apply_defaults(template):
        if template:
            settings, _ = PurchaseOrderSettings.objects.get_or_create(template_name=template)
            settings.porder_number_prefix = "P.ORD"
            settings.porder_starting_number = 1
            settings.header_background_color = "#ffa500"
            for key, value in default_labels.items():
                setattr(settings, key, value)
            settings.save()

    apply_defaults(template1)
    apply_defaults(template2)

    section = request.GET.get('section', 'purchase')
    return redirect(f'/user_purchase_order/?section={section}')

def get_payment_settings():
    settings = PaymentSettings.objects.first()
    if not settings:
        settings = PaymentSettings.objects.create()
    return settings

def payment_settings_view(request):
    settings = get_payment_settings()
    return render(request, 'admin/payment_settings.html', {'settings': settings})

def save_payment_settings(request):
    if request.method == 'POST':
        settings = get_payment_settings()

       
        settings.payment_receipt_label = request.POST.get('payment_receipt_label', 'Payment Receipt')
        settings.payment_for_invoice_label = request.POST.get('payment_for_invoice_label', 'Payment for Invoice #')
        settings.amount_received_from_label = request.POST.get('amount_received_from_label', 'Amount received from:')
        settings.description_label = request.POST.get('description_label', 'Description:')
        settings.payment_received_in_label = request.POST.get('payment_received_in_label', 'Payment Received in:')
        settings.payment_receipt_number_label = request.POST.get('payment_receipt_number_label', 'Payment Receipt #')
        settings.payment_date_label = request.POST.get('payment_date_label', 'Payment Date:')
        settings.payment_amount_label = request.POST.get('payment_amount_label', 'Payment Amount:')
        settings.total_amount_due_label = request.POST.get('total_amount_due_label', 'Total Amount Due')
        settings.total_paid_label = request.POST.get('total_paid_label', 'Total Paid')
        settings.balance_due_label = request.POST.get('balance_due_label', 'Balance Due')
        settings.payment_receipt_prefix = request.POST.get('payment_receipt_prefix', 'RCPT')

       
        settings.show_paid_on_fully_paid = 'show_paid_on_fully_paid' in request.POST
        settings.send_receipt_email_after_payment = 'send_receipt_email_after_payment' in request.POST
        settings.attach_invoice_to_receipt_email = 'attach_invoice_to_receipt_email' in request.POST

  
        restore_default_image = request.POST.get('restore_default_paid_image') == 'true'
        
        if restore_default_image:
          
            if settings.paid_image and default_storage.exists(settings.paid_image.name):
                default_storage.delete(settings.paid_image.name)
            settings.paid_image = None
        elif 'paid_image' in request.FILES:
         
            if settings.paid_image and default_storage.exists(settings.paid_image.name):
                default_storage.delete(settings.paid_image.name)
            settings.paid_image = request.FILES['paid_image']

        try:
            settings.save()
            messages.success(request, '')
        except Exception as e:
            messages.error(request, f'Error saving payment settings: {str(e)}')
        
        return redirect('payment_settings_view')  
    return redirect('payment_settings_view')

@csrf_exempt
def restore_paid_image_default(request):
    if request.method == 'POST':
        settings = get_payment_settings()
        if settings.paid_image and default_storage.exists(settings.paid_image.name):
            default_storage.delete(settings.paid_image.name)
        settings.paid_image = None
        settings.save()
        return JsonResponse({'success': True, 'message': 'Default PAID image restored'})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@csrf_exempt
def restore_payment_defaults(request):
    if request.method == 'POST':
        settings = get_payment_settings()

        defaults = {
            'payment_receipt_label': 'Payment Receipt',
            'payment_for_invoice_label': 'Payment for Invoice #',
            'amount_received_from_label': 'Amount received from:',
            'description_label': 'Description:',
            'payment_received_in_label': 'Payment Received in:',
            'payment_receipt_number_label': 'Payment Receipt #',
            'payment_date_label': 'Payment Date:',
            'payment_amount_label': 'Payment Amount:',
            'total_amount_due_label': 'Total Amount Due',
            'total_paid_label': 'Total Paid',
            'balance_due_label': 'Balance Due',
            'payment_receipt_prefix': 'RCPT',
        }

        for key, val in defaults.items():
            setattr(settings, key, val)

        settings.save()

        return JsonResponse({'success': True, 'message': 'Default payment labels restored', 'defaults': defaults})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
@require_http_methods(["POST"])
def save_email_template(request):
    try:
        template_name = request.POST.get('template_name')
        body = request.POST.get('body', '')
        html = request.POST.get('html', '')
        
        if not template_name:
            return JsonResponse({
                'success': False,
                'message': 'Template name is required'
            })
        
    
        email_template, created = EmailTemplate.objects.get_or_create(
            template_name=template_name,
            defaults={
                'body': body,
                'html': html
            }
        )
        
       
        if not created:
            email_template.body = body
            email_template.html = html
            email_template.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Template updated successfully',
                'template_id': email_template.id
            })
        else:
            return JsonResponse({
                'success': True,
                'message': 'Template created successfully',
                'template_id': email_template.id
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error saving template: {str(e)}'
        })

@login_required
@require_http_methods(["GET"])
def get_email_template(request):
    try:
        template_name = request.GET.get('template_name')
        
        if not template_name:
            return JsonResponse({
                'success': False,
                'message': 'Template name is required'
            })
        
        try:
            email_template = EmailTemplate.objects.get(template_name=template_name)
            
            return JsonResponse({
                'success': True,
                'data': {
                    'id': email_template.id,
                    'body': email_template.body,
                    'html': email_template.html,
                    'template_name': template_name
                }
            })
            
        except EmailTemplate.DoesNotExist:
            return JsonResponse({
                'success': True,
                'data': None,
                'message': 'No template found'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading template: {str(e)}'
        })

@login_required
def email_template_page(request):
    email_templates = EmailTemplate.objects.all()
    return render(request, 'admin/email_templates.html', {'email_templates': email_templates})

@login_required
def user_email_template_page(request):
    email_templates = EmailTemplate.objects.all()
    return render(request, 'user/user_email_templates.html', {'email_templates': email_templates})

@login_required
@require_http_methods(["POST"])
def user_save_email_template(request):
    try:
        template_name = request.POST.get('template_name')
        body = request.POST.get('body', '')
        html = request.POST.get('html', '')
        
        if not template_name:
            return JsonResponse({
                'success': False,
                'message': 'Template name is required'
            })
        
    
        email_template, created = EmailTemplate.objects.get_or_create(
            template_name=template_name,
            defaults={
                'body': body,
                'html': html
            }
        )
        
      
        if not created:
            email_template.body = body
            email_template.html = html
            email_template.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Template updated successfully',
                'template_id': email_template.id
            })
        else:
            return JsonResponse({
                'success': True,
                'message': 'Template created successfully',
                'template_id': email_template.id
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error saving template: {str(e)}'
        })

def user_payment_settings_view(request):
    settings = get_payment_settings()
    return render(request, 'user/user_payment_settings.html', {'settings': settings})

def user_save_payment_settings(request):
    if request.method == 'POST':
        settings = get_payment_settings()

       
        settings.payment_receipt_label = request.POST.get('payment_receipt_label', 'Payment Receipt')
        settings.payment_for_invoice_label = request.POST.get('payment_for_invoice_label', 'Payment for Invoice #')
        settings.amount_received_from_label = request.POST.get('amount_received_from_label', 'Amount received from:')
        settings.description_label = request.POST.get('description_label', 'Description:')
        settings.payment_received_in_label = request.POST.get('payment_received_in_label', 'Payment Received in:')
        settings.payment_receipt_number_label = request.POST.get('payment_receipt_number_label', 'Payment Receipt #')
        settings.payment_date_label = request.POST.get('payment_date_label', 'Payment Date:')
        settings.payment_amount_label = request.POST.get('payment_amount_label', 'Payment Amount:')
        settings.total_amount_due_label = request.POST.get('total_amount_due_label', 'Total Amount Due')
        settings.total_paid_label = request.POST.get('total_paid_label', 'Total Paid')
        settings.balance_due_label = request.POST.get('balance_due_label', 'Balance Due')
        settings.payment_receipt_prefix = request.POST.get('payment_receipt_prefix', 'RCPT')

       
        settings.show_paid_on_fully_paid = 'show_paid_on_fully_paid' in request.POST
        settings.send_receipt_email_after_payment = 'send_receipt_email_after_payment' in request.POST
        settings.attach_invoice_to_receipt_email = 'attach_invoice_to_receipt_email' in request.POST

  
        restore_default_image = request.POST.get('restore_default_paid_image') == 'true'
        
        if restore_default_image:
          
            if settings.paid_image and default_storage.exists(settings.paid_image.name):
                default_storage.delete(settings.paid_image.name)
            settings.paid_image = None
        elif 'paid_image' in request.FILES:
         
            if settings.paid_image and default_storage.exists(settings.paid_image.name):
                default_storage.delete(settings.paid_image.name)
            settings.paid_image = request.FILES['paid_image']

        try:
            settings.save()
            messages.success(request, '')
        except Exception as e:
            messages.error(request, f'Error saving payment settings: {str(e)}')
        
        return redirect('user_payment_settings_view')  
    return redirect('user_payment_settings_view')




from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils.decorators import method_decorator
from django.views import View
from django.db.models import Sum
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from io import BytesIO
import base64
import mimetypes
import json
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import json
import base64
from io import BytesIO
from django.views.decorators.http import require_POST
from django.core.mail import EmailMessage
from django.conf import settings
from django.shortcuts import get_object_or_404
from django.db.models import Sum
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl.styles import Font
from .models import Customer, Invoice, Companies 



@login_required
def get_customer(request, customer_id):
    """
    API endpoint to get customer data for editing
    """
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        
        customer_data = {
            'id': customer.id,
            'customer_id': customer.customer_id,
            'category_name': customer.customer_category.category if customer.customer_category else '',
            'category_id': customer.customer_category.id if customer.customer_category else '',
            'status': customer.status,
            'business_name': customer.business_name,
            'address': customer.address,
            'ship_to_name': customer.ship_to_name,
            'shipping_address': customer.shipping_address,
            'contact_person': customer.contact_person,
            'email_address': customer.email_address,
            'telephone_number': customer.telephone_number,
            'fax_number': customer.fax_number,
            'sms_mobile_number': customer.sms_mobile_number,
            'ship_to_contact_person': customer.ship_to_contact_person,
            'ship_to_email_address': customer.ship_to_email_address,
            'ship_to_telephone_number': customer.ship_to_telephone_number,
            'ship_to_fax_number': customer.ship_to_fax_number,
            'tax_exempt': customer.tax_exempt,
            'specific_tax1_percent': customer.specific_tax1_percent,
            'specific_tax2_percent': customer.specific_tax2_percent,
            'discount_percent': customer.discount_percent,
            'customer_type': customer.customer_type,
            'country': customer.country,
            'city': customer.city,
            'notes': customer.notes,
        }
        
        return JsonResponse({
            'success': True,
            'customer': customer_data
        })
    
    except Customer.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Customer not found'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def edit_customer(request, customer_id):
    """
    Handle customer edit form submission
    """
    if request.method == 'POST':
        try:
            customer = get_object_or_404(Customer, id=customer_id)
            
           
            category = None
            category_id = request.POST.get('customer_category')
            category_name = request.POST.get('category_name', '').strip()

            
            if category_id:
                category = get_object_or_404(CustomerCategory, id=category_id)
            elif category_name:
                category, created = CustomerCategory.objects.get_or_create(
                    category=category_name
                )
                if created:
                    messages.success(request, f'New category "{category_name}" created.')
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Please select or enter a category.'
                })

            
            customer.customer_id = request.POST.get('customer_id')
            customer.customer_category = category
            customer.status = 'Active' if request.POST.get('status') else 'Inactive'
            customer.business_name = request.POST.get('business_name')
            customer.address = request.POST.get('address')
            customer.ship_to_name = request.POST.get('ship_to_name', '')
            customer.shipping_address = request.POST.get('shipping_address', '')
            customer.contact_person = request.POST.get('contact_person')
            customer.email_address = request.POST.get('email_address')
            customer.telephone_number = request.POST.get('telephone_number')
            customer.fax_number = request.POST.get('fax_number', '')
            customer.sms_mobile_number = request.POST.get('sms_mobile_number', '')
            customer.ship_to_contact_person = request.POST.get('ship_to_contact_person', '')
            customer.ship_to_email_address = request.POST.get('ship_to_email_address', '')
            customer.ship_to_telephone_number = request.POST.get('ship_to_telephone_number', '')
            customer.ship_to_fax_number = request.POST.get('ship_to_fax_number', '')
            customer.tax_exempt = bool(request.POST.get('tax_exempt'))
            customer.specific_tax1_percent = float(request.POST.get('specific_tax1_percent', 0) or 0)
            customer.specific_tax2_percent = float(request.POST.get('specific_tax2_percent', 0) or 0)
            customer.discount_percent = float(request.POST.get('discount_percent', 0) or 0)
            customer.customer_type = request.POST.get('customer_type', 'client')
            customer.country = request.POST.get('country')
            customer.city = request.POST.get('city')
            customer.notes = request.POST.get('notes', '')

            customer.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Customer updated successfully'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


@login_required
def get_tax_settings(request):
    """
    API endpoint to get current tax settings for dynamic customer form
    """
    try:
       
        company = Companies.objects.first()
        
        if company:
           
            tax = Tax.objects.filter(company=company).first()
            
            if tax:
                tax_type = tax.tax_type_first
            else:
                tax_type = 'no'  
        else:
            tax_type = 'no'  
        
        return JsonResponse({
            'tax_type': tax_type,
            'success': True
        })
    
    except Exception as e:
        return JsonResponse({
            'tax_type': 'no',
            'success': False,
            'error': str(e)
        })



@login_required
def customer_list(request):
    """
    Display customer list with categories and user permissions
    """

    
    categories = CustomerCategory.objects.all()

   
    category_filter = request.GET.get('category', 'all')

   
    base_query = Customer.objects.all()

    
    if category_filter == 'all' or category_filter == 'default':
        customers = base_query
    elif category_filter == 'client':
        customers = base_query.filter(customer_type='client')
    elif category_filter == 'vendor':
        customers = base_query.filter(customer_type='vendor')
    elif category_filter == 'client_vendor':
        customers = base_query.filter(customer_type='both')
    else:
        customers = base_query.filter(
            customer_category__category__iexact=category_filter
        )

   
    company = Companies.objects.first()

    
    static_categories = ['client_vendor', 'client', 'vendor', 'default']
    dynamic_categories = CustomerCategory.objects.exclude(
        category__in=static_categories
    ).values_list('category', flat=True).distinct()

    
    try:
        modules = Modules.objects.get(user=request.user)
    except Modules.DoesNotExist:
        modules = None

    context = {
        'categories': categories,
        'customers': customers,
        'current_filter': category_filter,
        'company': company,
        'dynamic_categories': dynamic_categories,
        'user_permissions': modules  
    }

    
    if request.user.is_superuser:
        return render(request, 'admin/customer_list.html', context)
    else:
        return render(request, 'user/customer_list.html', context)



@login_required
def add_customer(request):
    """
    Handle customer creation
    """
    if request.method == 'POST':
        try:
            category = None
            category_id = request.POST.get('customer_category')
            category_name = request.POST.get('category_name', '').strip()

           
            if category_id:
                category = get_object_or_404(CustomerCategory, id=category_id)
            elif category_name:
                category, created = CustomerCategory.objects.get_or_create(
                    category=category_name
                )
                # if created:
                #     messages.success(request, f'New category "{category_name}" created.')
            else:
               
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': 'Please select or enter a category.'
                    })
                messages.error(request, 'Please select or enter a category.')
                return redirect('customer_list')

           
            customer = Customer(
                user=request.user,
                customer_id=request.POST.get('customer_id'),
                customer_category=category,
                status='Active' if request.POST.get('status') else 'Inactive',
                business_name=request.POST.get('business_name'),
                address=request.POST.get('address'),
                ship_to_name=request.POST.get('ship_to_name', ''),
                shipping_address=request.POST.get('shipping_address', ''),
                contact_person=request.POST.get('contact_person'),
                email_address=request.POST.get('email_address'),
                telephone_number=request.POST.get('telephone_number'),
                fax_number=request.POST.get('fax_number', ''),
                sms_mobile_number=request.POST.get('sms_mobile_number', ''),
                ship_to_contact_person=request.POST.get('ship_to_contact_person', ''),
                ship_to_email_address=request.POST.get('ship_to_email_address', ''),
                ship_to_telephone_number=request.POST.get('ship_to_telephone_number', ''),
                ship_to_fax_number=request.POST.get('ship_to_fax_number', ''),
                tax_exempt=bool(request.POST.get('tax_exempt')),
                specific_tax1_percent=float(request.POST.get('specific_tax1_percent', 0) or 0),
                specific_tax2_percent=float(request.POST.get('specific_tax2_percent', 0) or 0),
                discount_percent=float(request.POST.get('discount_percent', 0) or 0),
                customer_type=request.POST.get('customer_type', 'client'),
                country=request.POST.get('country'),
                city=request.POST.get('city'),
                notes=request.POST.get('notes', ''),
            )

            customer.save()
            
           
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Customer saved successfully!'
                })
            
           
           
        except Exception as e:
           
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': f'Error adding customer: {str(e)}'
                })
            messages.error(request, f'Error adding customer: {str(e)}')

    return redirect('customer_list')


@login_required
def get_categories(request):
    """
    API endpoint to get all categories for dynamic loading
    """
    try:
        categories = CustomerCategory.objects.all().values('id', 'category')
        return JsonResponse({
            'categories': list(categories),
            'success': True
        })
    except Exception as e:
        return JsonResponse({
            'categories': [],
            'success': False,
            'error': str(e)
        })





@csrf_exempt
def delete_customers(request):
    if request.method == "POST":
        data = json.loads(request.body)
        ids = data.get('customer_ids', [])
        Customer.objects.filter(id__in=ids).delete()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)



def preview_invoice(request):
    customer_code = request.GET.get('customer_id')
    customer = get_object_or_404(Customer, customer_id=customer_code)

    invoices = Invoice.objects.filter(invoice_to=customer) 
    total = sum(inv.invoice_total for inv in invoices)
    paid = sum(inv.total_paid for inv in invoices)
    balance = sum(inv.balance for inv in invoices)

    company = Companies.objects.first()

    context = {
        'customer': customer,
        'invoices': invoices,
        'total': total,
        'paid': paid,
        'balance': balance,
        'company': company,  
    }
    return render(request, 'admin/invoice_preview.html', context)





class EmailInvoiceView(View):
    """
    Handle email invoice sending functionality
    """
    
    def get(self, request):
        """
        Render the email invoice modal (if needed as separate page)
        """
        return render(request, 'email_invoice_modal.html')
    
    @method_decorator(csrf_exempt)
    def post(self, request):
        """
        Handle email sending request
        """
        try:
           
            data = json.loads(request.body)
            
           
            email_to = data.get('email_to')
            carbon_copy = data.get('carbon_copy', '')
            subject = data.get('subject', 'Invoice List')
            message = data.get('message', '')
            customer_id = data.get('customer_id')
            business_name = data.get('business_name', 'Customer')
            attachments = data.get('attachments', [])
            
          
            if not email_to:
                return JsonResponse({'success': False, 'error': 'Email address is required'})
            
            if not subject:
                return JsonResponse({'success': False, 'error': 'Subject is required'})
            
            if not customer_id:
                return JsonResponse({'success': False, 'error': 'Customer ID is required'})
            
           
            result = self.send_invoice_email(
                email_to=email_to,
                carbon_copy=carbon_copy,
                subject=subject,
                message=message,
                customer_id=customer_id,
                business_name=business_name,
                attachments=attachments
            )
            
            return JsonResponse(result)
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    

    
    def generate_invoice_pdf(self, customer_id, business_name):
        """
        Generate invoice PDF with real data from database
        """
        try:
           
            customer = get_object_or_404(Customer, customer_id=customer_id)
            
            
            try:
                company = Companies.objects.first()  
            except:
                company = None
            
            
            invoices = Invoice.objects.filter(customer=customer).order_by('-invoice_date')
            
           
            total_invoice_amount = invoices.aggregate(total=Sum('invoice_total'))['total'] or 0
            total_paid_amount = invoices.aggregate(total=Sum('total_paid'))['total'] or 0
            total_balance = total_invoice_amount - total_paid_amount
            
            buffer = BytesIO()
            
           
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            story = []
            
          
            styles = getSampleStyleSheet()
            title_style = styles['Title']
            normal_style = styles['Normal']
            heading_style = styles['Heading2']
            

            if company:
                company_info = f"""
                <b>{company.company_name}</b><br/>
                {company.address}<br/>
                Email: {company.email}<br/>
                Sales tax reg. No: {company.sales_tax_reg_no}
                """
                company_para = Paragraph(company_info, normal_style)
                story.append(company_para)
                story.append(Spacer(1, 20))
            
           
            title = Paragraph(f"Invoice List - Customer ID: {customer.customer_id}", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
           
            customer_info = f"""
            <b>Bill to:</b><br/>
            <b>{customer.business_name}</b><br/>
            {customer.address}<br/><br/>
            
            <b>Ship to:</b><br/>
            <b>{customer.ship_to_name or customer.business_name}</b><br/>
            {customer.shipping_address or customer.address}<br/>
            """
            customer_para = Paragraph(customer_info, normal_style)
            story.append(customer_para)
            story.append(Spacer(1, 20))
            

            invoice_data = [
                ['Invoice No.', 'Date', 'Due Date', 'Recurring', 'Status', 'Invoice Total', 'Total Paid', 'Balance']
            ]
            
            for invoice in invoices:
                due_date = invoice.due_date.strftime('%Y-%m-%d') if invoice.due_check and invoice.due_date else '-'
                recurring = 'Yes' if invoice.recurring else 'No'
                
                invoice_data.append([
                    invoice.invoice_number,
                    invoice.invoice_date.strftime('%Y-%m-%d'),
                    due_date,
                    recurring,
                    invoice.status,
                    f'₹ {invoice.invoice_total:,.2f}',
                    f'₹ {invoice.total_paid:,.2f}',
                    f'₹ {invoice.balance:,.2f}'
                ])
            
           
            invoice_data.append([
                '- End of List - Summary:',
                '', '', '', '',
                f'₹ {total_invoice_amount:,.2f}',
                f'₹ {total_paid_amount:,.2f}',
                f'₹ {total_balance:,.2f}'
            ])
            
          
            table = Table(invoice_data, colWidths=[0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
               
                ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -2), 9),
                
               
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 10),
                
               
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                
               
                ('ALIGN', (5, 0), (-1, -1), 'RIGHT'),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 30))
            
          
            footer = Paragraph("Thank you for your business!", normal_style)
            story.append(footer)
            
           
            doc.build(story)
            
          
            pdf_content = buffer.getvalue()
            buffer.close()
            
            return pdf_content
            
        except Customer.DoesNotExist:
            raise Exception(f"Customer with ID {customer_id} not found")
        except Exception as e:
            raise Exception(f"Error generating PDF: {str(e)}")





@require_POST
@csrf_exempt
def send_email_invoice(request):
    """
    Enhanced function-based view for sending email invoices with conditional PDF generation
    """
    try:
        
        data = json.loads(request.body)
        
        email_to = data.get('email_to')
        carbon_copy = data.get('carbon_copy', '')
        subject = data.get('subject', 'Invoice List')
        message = data.get('message', '')
        customer_id = data.get('customer_id')
        business_name = data.get('business_name', 'Customer')
        attachments = data.get('attachments', []) 
        include_auto_pdf = data.get('include_auto_pdf', True)  
        auto_pdf_data = data.get('auto_pdf_data')  
        
       
        if not email_to:
            return JsonResponse({'success': False, 'error': 'Email address is required'})
        
        if not customer_id:
            return JsonResponse({'success': False, 'error': 'Customer ID is required'})
        
        
        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email_to],
            cc=[carbon_copy] if carbon_copy else [],
        )
        
        
        if '<' in message and '>' in message:
            email.content_subtype = 'html'
        
        attachment_count = 0
        
       
        if include_auto_pdf and auto_pdf_data:
            try:
                pdf_content = generate_invoice_pdf_with_real_data(
                    customer_id, 
                    auto_pdf_data.get('businessName', business_name)
                )
                pdf_filename = auto_pdf_data.get('filename', f'{business_name}_Invoices_List.pdf')
                email.attach(pdf_filename, pdf_content, 'application/pdf')
                attachment_count += 1
                print(f"Attached auto-generated PDF: {pdf_filename}")
            except Exception as e:
                print(f"Error generating auto PDF: {str(e)}")
               
                pass
        
       
        for attachment in attachments:
            if 'file_data' in attachment and 'filename' in attachment:
                try:
                    
                    file_data = base64.b64decode(attachment['file_data'])
                    content_type = attachment.get('content_type', 'application/octet-stream')
                    filename = attachment['filename']
                    
                   
                    email.attach(filename, file_data, content_type)
                    attachment_count += 1
                    print(f"Attached additional file: {filename}")
                    
                except Exception as e:
                    print(f"Error processing attachment {attachment.get('filename', 'unknown')}: {str(e)}")
                   
                    continue
        
       
        email.send()
        
        return JsonResponse({
            'success': True, 
            'message': f'Email sent successfully with {attachment_count} attachment{"s" if attachment_count != 1 else ""}!',
            'attachment_count': attachment_count
        })
        
    except Exception as e:
        print(f"Email sending error: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

def generate_invoice_pdf_with_real_data(customer_id, business_name):
    """
    Generate PDF invoice with real data using reportlab canvas
    """
    try:
        
        customer = get_object_or_404(Customer, customer_id=customer_id)
        
       
        try:
            company = Companies.objects.first()
        except:
            company = None
        
      
        invoices = Invoice.objects.filter(invoice_to=customer).order_by('-invoice_date')
        
       
        total_invoice_amount = invoices.aggregate(total=Sum('invoice_total'))['total'] or 0
        total_paid_amount = invoices.aggregate(total=Sum('total_paid'))['total'] or 0
        total_balance = total_invoice_amount - total_paid_amount
        
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
       
        width, height = letter
        
       
        y_position = height - 50
        if company:
            p.setFont("Helvetica-Bold", 14)
            p.drawString(50, y_position, company.company_name)
            y_position -= 20
            
            p.setFont("Helvetica", 10)
            p.drawString(50, y_position, company.address)
            y_position -= 15
            p.drawString(50, y_position, f"Email: {company.email}")
            y_position -= 15
            p.drawString(50, y_position, f"Sales tax reg. No: {company.sales_tax_reg_no}")
            y_position -= 30
        
        
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, y_position, "Invoice List")
        p.drawString(400, y_position, f"Customer ID: {customer.customer_id}")
        y_position -= 40
        
       
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y_position, "Bill to:")
        p.drawString(300, y_position, "Ship to:")
        y_position -= 20
        
        p.setFont("Helvetica", 10)
        p.drawString(50, y_position, customer.business_name)
        p.drawString(300, y_position, customer.ship_to_name or customer.business_name)
        y_position -= 15
        p.drawString(50, y_position, customer.address)
        p.drawString(300, y_position, customer.shipping_address or customer.address)
        y_position -= 40
        
     
        p.setFont("Helvetica-Bold", 9)
        headers = ["Invoice No.", "Date", "Due Date", "Recurring", "Status", "Invoice Total", "Total Paid", "Balance"]
        x_positions = [50, 120, 180, 230, 280, 330, 420, 500]
        
        for i, header in enumerate(headers):
            p.drawString(x_positions[i], y_position, header)
        
        y_position -= 20
        
      
        p.setFont("Helvetica", 8)
        for invoice in invoices:
            if y_position < 100:  
                p.showPage()
                y_position = height - 50
            
            due_date = invoice.due_date.strftime('%Y-%m-%d') if invoice.due_check and invoice.due_date else '-'
            recurring = 'Yes' if invoice.recurring else 'No'
            
            data = [
                invoice.invoice_number,
                invoice.invoice_date.strftime('%Y-%m-%d'),
                due_date,
                recurring,
                invoice.status,
                f'₹ {invoice.invoice_total:,.2f}',
                f'₹ {invoice.total_paid:,.2f}',
                f'₹ {invoice.balance:,.2f}'
            ]
            
            for i, value in enumerate(data):
                p.drawString(x_positions[i], y_position, str(value))
            
            y_position -= 15
        
    
        y_position -= 10
        p.setFont("Helvetica-Bold", 9)
        p.drawString(50, y_position, "- End of List - Summary:")
        p.drawString(330, y_position, f'₹ {total_invoice_amount:,.2f}')
        p.drawString(420, y_position, f'₹ {total_paid_amount:,.2f}')
        p.drawString(500, y_position, f'₹ {total_balance:,.2f}')
        
      
        p.setFont("Helvetica-Oblique", 10)
        p.drawString(50, 50, "Thank you for your business!")
        
        p.save()
        
        pdf_content = buffer.getvalue()
        buffer.close()
        
        return pdf_content
        
    except Customer.DoesNotExist:
        raise Exception(f"Customer with ID {customer_id} not found")
    except Exception as e:
        raise Exception(f"Error generating PDF: {str(e)}")




def export_customers_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

   
    headers = [
        "Customer ID", "Category", "Business Name", "Contact Person",
        "Email", "Telephone", "SMS Number", "Customer Type",
        "Country", "City", "Status", "Shipping Address"
    ]
    for col_num, header in enumerate(headers, 1):
        ws[f'{get_column_letter(col_num)}1'] = header

    # Data rows
    for row_num, customer in enumerate(Customer.objects.all(), 2):
        row = [
            customer.customer_id,
            customer.customer_category.category,
            customer.business_name,
            customer.contact_person,
            customer.email_address,
            customer.telephone_number,
            customer.sms_mobile_number,
            customer.get_customer_type_display(),
            customer.country,
            customer.city,
            customer.status,
            customer.shipping_address,
        ]
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num).value = value

   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=customers.xlsx'
    wb.save(response)
    return response



def download_excel_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CustomerTemplate"

    headers = [
        "CUSTOMER ID", "CUSTOMER NAME", "CATEGORY", "ADDRESS", "TEL", "FAX", "EMAIL",
        "CONTACT PERSON", "SHIP TO NAME", "SHIP TO ADDRESS", "SHIP TO TEL", "SHIP TO FAX",
        "DISCOUNT", "SPECIAL TAX 1", "SPECIAL TAX 2", "ACTIVE", "TAX EXEMPTED"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_template.xlsx'
    wb.save(response)
    return response


@csrf_exempt
def import_customers_from_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        def clean_text(value):
            """Convert numbers to int string to avoid decimal point, and clean up text"""
            if isinstance(value, float):
                if value.is_integer():
                    return str(int(value))
                return str(value)
            return str(value).strip() if value else ""

        count = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if not any(row):
                continue  

            try:
                customer_id = clean_text(row[0])
                name = clean_text(row[1])
                category = clean_text(row[2])
                address = clean_text(row[3])
                tel = clean_text(row[4])
                fax = clean_text(row[5])
                email = clean_text(row[6])
                contact_person = clean_text(row[7])
                ship_to_name = clean_text(row[8])
                ship_to_address = clean_text(row[9])
                ship_tel = clean_text(row[10])
                ship_fax = clean_text(row[11])
                discount = float(row[12]) if row[12] is not None else 0.0
                tax1 = float(row[13]) if row[13] is not None else 0.0
                tax2 = float(row[14]) if row[14] is not None else 0.0
                active = clean_text(row[15]).lower()
                tax_exempt = clean_text(row[16]).lower()

                category_obj, _ = CustomerCategory.objects.get_or_create(category=category)

                Customer.objects.create(
                    user=request.user, 
                    customer_id=customer_id,
                    customer_category=category_obj,
                    status='Active' if active == 'active' else 'Inactive',
                    business_name=name,
                    address=address,
                    telephone_number=tel,
                    fax_number=fax,
                    email_address=email,
                    contact_person=contact_person,
                    ship_to_name=ship_to_name,
                    shipping_address=ship_to_address,
                    ship_to_telephone_number=ship_tel,
                    ship_to_fax_number=ship_fax,
                    discount_percent=discount,
                    specific_tax1_percent=tax1,
                    specific_tax2_percent=tax2,
                    tax_exempt=True if tax_exempt in ['yes', 'true', '1'] else False,
                    sms_mobile_number=tel,
                    ship_to_contact_person=contact_person,
                    ship_to_email_address=email,
                    customer_type="client",
                    country="Unknown",
                    city="Unknown",
                    notes=""
                )
                count += 1

            except Exception as e:
                print(f"Row {idx + 2} failed: {e}")
                continue

        return JsonResponse({"status": "success", "imported": count})
    
    return JsonResponse({"status": "error", "message": "Invalid request"})
